from openpyxl import Workbook, load_workbook
from tables_and_tuples import *
from functions import lambda11

# Libraries for plotting
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
from openpyxl.drawing.image import Image
import matplotlib.patches as patches

def open_workbook(path='', file_name='temp'):
    if file_name == 'temp':
        wb = Workbook()
    else:
        try:
            wb = load_workbook(path + file_name + '.xlsx')
        except FileNotFoundError:
            wb = Workbook()
    return wb

def write_to_excel(wb, cn, sheet_name='sheet', time=0, fun_evals=0, evals=0,
                   path='C:\\Users\LaHie1\Documents\Dippa\TEMP\\'):

    ws = wb.create_sheet(title=sheet_name)

    ws.append([])
    ws.append(['Elapsed time: ',time,'[s]'])
    ws.append(['Function evaluations: ', str(fun_evals)])
    ws.append(['Total evaluations: ', str(evals)])
    ws.append([])

    ws.append(['Connection configuration = ', cn.joint_config, cn.column_config, cn.sway])
    ws.append(['Global structure analysis method = ', cn.analysis_method])
    ws.append([])
    if cn.column.size != "":
        ws.append(["Column: ", cn.column.size])
        ws.append(['h = ', cn.column.h,' [mm]', 'b = ', cn.column.b, ' [mm]', 't_f = ', cn.column.t_f, ' [mm]',
                   't_w = ', cn.column.t_w, ' [mm]', 'r = ', cn.column.r, ' [mm]'])

        ws.append(['Material: ', cn.column.material, '', 'f_y = ', cn.column.f_y, ' [MPa]', 'f_u = ',
                   cn.column.f_u, ' [MPa]'])

        ws.append(['Cross-section properties:'])
        ws.append(['A = ', cn.column.A, ' [mm^2]', 'A_v = ', cn.column.A_v, ' [mm^2]', 'd_w = ', cn.column.d_w,' [mm]'])
        ws.append(['W_el_y = ', cn.column.W_el_y, ' [mm^3]', 'W_el_z = ', cn.column.W_el_z, ' [mm^3]'])
        ws.append(['W_pl_y = ', cn.column.W_pl_y, ' [mm^3]', 'W_pl_z = ', cn.column.W_pl_z, ' [mm^3]'])

        ws.append(['Cross-section class: ', cn.column.PL])
        ws.append(['Loads: '])
        ws.append(['N = ', str(round(cn.column.N, 5)), '[N]', 'Mt = ', str(round(cn.column.Mt, 5)), '[kNm]'])
        ws.append(['Vy = ', str(round(cn.column.Vy, 5)), '[N]', 'My = ', str(round(cn.column.My, 5)), '[kNm]'])
        ws.append(['Vz = ', str(round(cn.column.Vz, 5)), '[N]', 'Mz = ', str(round(cn.column.Mz, 5)), '[kNm]'])
        ws.append([])
        ws.append([])

    for beam_id in range(0, 2):

        sd = cn.side[beam_id]

        if len(sd.beam.size) != 0:
            ws.append([])
            ws.append(['V_wp_Rd = ', cn.V_wp_Rd * 1.0e-3, ' [kN]', ' (Column web panel in shear)'])
            ws.append(['F_c_wc_Rd = ', sd.F_c_wc_Rd * 1.0e-3, ' [kN]', ' (Column web in transverse compression)'])
            ws.append(['F_c_fb_Rd = ', sd.F_c_fb_Rd * 1.0e-3, ' [kN]', ' (Beam flange and web in compression)'])
            ws.append(['F_comp_Rd = ', sd.F_comp_Rd * 1.0e-3, ' [kN]',
                       ' (Design resistance of compression side, beta = ' + str(sd.beta)])
            ws.append([])

            ws.append(['Effective spring stiffness of components on compression side:'])
            ws.append(['k_1 = ', sd.k_1, ' [mm]', ' (Column web panel in shear)'])
            ws.append(['k_2 = ', sd.k_2, ' [mm]', ' (Column web in compression)'])
            ws.append(
                ['k_eff_comp = ', sd.k_eff_comp, ' [mm]', ' (Total effective spring stiffness of compression side)'])
            ws.append(['Effective spring stiffness of components on tension side:'])
            ws.append(['k_eq = ', sd.k_eq, ' [mm]', ' (Total effective spring stiffness of tension components)'])
            ws.append(['z_eq = ', sd.z_eq, ' [mm]', ' (Total effective lever arm of tension components)'])
            ws.append([])

            ws.append(['Connection strength classification: ', sd.strength_class])
            ws.append(['Moment capacity of endplate connection M_j_Rd = ', sd.M_j_Rd * 1.0e-6, ' [kNm]'])
            ws.append(['(M_rigid = ', sd.M_rigid * 1.0e-6, ' [kNm]', 'M_pinned = ', sd.M_pinned * 1.0e-6, ' [kNm]'])
            ws.append([])
            ws.append(['Connection stiffness classification: ', sd.stiffness_class])
            ws.append(['Rotational stiffness of endplate connection S_j_ini = ', sd.S_j_ini * 1.0e-6, ' [kNm/rad]'])
            ws.append(['Rotational stiffness of endplate connection S_j = ', sd.S_j * 1.0e-6, ' [kNm/rad]'])
            ws.append(['(S_j_rigid = ', sd.S_j_rigid * 1.0e-6, ' [kNm/rad], S_j_pinned = ',
                       sd.S_j_pinned * 1.0e-6, ' [kNm/rad])'])
            ws.append([])
            ws.append([])

            # if cost_info:
            ws.append(['\nUsed cost function: ', cn.cost_function])
            ws.append(['Total costs on side ' + str(beam_id) + ': ', sd.cost, ' [e]'])
            ws.append(['Material costs:'])
            ws.append(['Plate: ', sd.material_cost.plate, ' [e]'])
            ws.append(['Bolts: ', sd.material_cost.bolts, ' [e]'])
            ws.append(['Work step costs:'])
            ws.append(['Blasting: ', sd.blasting.cost, ' [e]'])
            ws.append(['Cutting: ', sd.cutting.cost, ' [e]'])
            # ws.append(['Drilling: ',sd.drilling.cost,' [e]'])
            ws.append(['Part assembly: ', sd.part_assembly.cost, ' [e]'])
            ws.append(['Post treatment: ', sd.post_treatment.cost, ' [e]'])
            ws.append(['Painting: ', sd.painting.cost, ' [e]'])
            ws.append([])

            ws.append(['\nConnection utilization ratios:'])
            ws.append(['Side ' + str(beam_id)])
            ws.append(['   M_Ed/M_Rd = ', sd.n_M, ' (Connection moment utilization rate)'])
            ws.append(['   V_Ed/V_Rd = ', sd.n_V, ' (Connection shear utilization rate)'])
            ws.append([])

            ws.append(["Beam: ", sd.beam.size, ', side ' + str(beam_id)])
            ws.append(['h = ', sd.beam.h, ' [mm]', 'b = ', sd.beam.b, ' [mm]', 't_f = ', sd.beam.t_f, ' [mm]',
                       't_w = ', sd.beam.t_w, ' [mm]', 'r = ', sd.beam.r, ' [mm]'])

            ws.append(['Material: ', sd.beam.material, '', 'f_y = ', sd.beam.f_y, ' [MPa]', 'f_u = ',
                       sd.beam.f_u, ' [MPa]'])

            ws.append(['Cross-section properties:'])
            ws.append(
                ['A = ', sd.beam.A, ' [mm^2]', 'A_v = ', sd.beam.A_v, ' [mm^2]', 'd_w = ', sd.beam.d_w, ' [mm]'])
            ws.append(['W_el_y = ', sd.beam.W_el_y, ' [mm^3]', 'W_el_z = ', sd.beam.W_el_z, ' [mm^3]'])
            ws.append(['W_pl_y = ', sd.beam.W_pl_y, ' [mm^3]', 'W_pl_z = ', sd.beam.W_pl_z, ' [mm^3]'])

            ws.append(['Cross-section class: ', sd.beam.PL])
            ws.append(['Loads: '])
            ws.append(['N = ', str(round(sd.beam.N, 5)), '[N]', 'Mt = ', str(round(sd.beam.Mt, 5)), '[kNm]'])
            ws.append(['Vy = ', str(round(sd.beam.Vy, 5)), '[N]', 'My = ', str(round(sd.beam.My, 5)), '[kNm]'])
            ws.append(['Vz = ', str(round(sd.beam.Vz, 5)), '[N]', 'Mz = ', str(round(sd.beam.Mz, 5)), '[kNm]'])
            ws.append([])
            ws.append([])

            ws.append(['Plate information:'])
            ws.append(['h_p = ', sd.plate_height, '[mm]', 'b_p = ', sd.plate_width, '[mm]',
                       't_p = ', sd.plate_thickness, '[mm]',
                       'Upper overhang = ', sd.upper_overhang, '[mm]', 'Lower overhang = ', sd.lower_overhang, '[mm]'])
            ws.append(['Material: ', sd.plate_material, '', 'f_y = ', mat[sd.plate_material]['f_y'], ' [MPa]', 'f_u = ',
                       mat[sd.plate_material]['f_u'], ' [MPa]'])
            ws.append([])

            ws.append([sd.welds])
            ws.append(['a_f = ', sd.weld_f, ' mm', 'a_w = ', sd.weld_w, ' mm'])
            ws.append([])
            ws.append(["Compression center " + str(sd.compression_center) + " [mm] from top of the end-plate"])
            ws.append([])
            ws.append([])

            for j in range(0, len(sd.row)):
                row = sd.row[j]

                ws.append(['Row ', row.id+1,'z = ', row.z,'Bolt = ',row.bolt])
                ws.append(['Location on flange side = ',row.location_f])
                ws.append(['Location on plate side = ',row.location_p])
                ws.append(['w = ', row.w, '[mm]', ' h_r = ', row.h_r, '[mm]'])
                ws.append(['e = ',row.e, '[mm]','m = ',row.m, '[mm]'])
                ws.append(['ec = ',row.ec, '[mm]','mc = ',row.mc, '[mm]'])
                ws.append(['e_min = ', row.e_min, '[mm]'])
                ws.append(['e_x = ',row.e_x, '[mm]','m_x = ',row.m_x, '[mm]'])
                ws.append(['e_1 = ', row.e_1, '[mm]'])
                ws.append(['L = ',row.L, '[mm]',', L_b = ',row.L_b, '[mm]'])
                ws.append(['l_eff_1f = ',row.l_eff_1f, '[mm]',' mm, l_eff_2f = ',row.l_eff_2f, '[mm]'])
                ws.append(['l_eff_1p = ',row.l_eff_1p, '[mm]',' mm, l_eff_2p = ',row.l_eff_2p, '[mm]'])
                ws.append(['l_eff_cpf [mm]'])
                ws.append(row.l_eff_cpf)
                ws.append(['l_eff_ncf [mm]'])
                ws.append(row.l_eff_ncf)
                ws.append(['l_eff_cpp [mm]'])
                ws.append(row.l_eff_cpp)
                ws.append(['l_eff_ncp [mm]'])
                ws.append(row.l_eff_ncp)

                if row.location_p == "First bolt-row below tension flange of beam":
                    ws.append(['lambda1 = ',row.lambda1,'lambda2 = ',row.lambda2])
                    ws.append(['alpha = ',row.alfa])

                    #plt.figure("alfa")
                    #plt.xlabel('$\\lambda$1')
                    #plt.ylabel('$\\lambda$2')

                    #for i in range(0, 280):
                    #    plt.scatter(lambda11(i / 200.0, row.alfa), i / 200.0, s=1.0, c="r")
                    #plt.scatter(row.lambda1, row.lambda2, marker="o", s=5.0, c="b")

                    # For image with alfa values shown
                    #img = mpimg.imread('plot_alfa_2.png')
                    #plt.imshow(img, extent=[0.0, 1.053, 0.0, 1.59])

                    # Saving alfa plot
                    #plt.savefig(path + 'alfa_plot.png', dpi=150)
                    #plt.close()
                    #img2 = Image(path + 'alfa_plot.png')
                    #ws.add_image(img2, "K125")

                ws.append(['Total effective spring stiffness of row is k_eff = ', row.k_eff,' [mm]'])
                ws.append(['k_3 = ',row.k_3,' [mm]',' (Column web tension)'])
                ws.append(['k_4 = ',row.k_4,' [mm]',' (Column flange bending)'])
                ws.append(['k_5 = ',row.k_5,' [mm]',' (End-plate bending)'])
                ws.append(['k_10 = ',row.k_10,' [mm]',' (Bolts)'])

                ws.append(['Shear capacity of row is ',row.V_Rd*1.0e-3,' [kN]'])
                ws.append(['Tension capacity of row is defined by ', row.mode])
                ws.append(['Total tension capacity is ',row.F_t_Rd*1.0e-3,' [kN]'])

                ws.append(['F_t_wc_Rd = ',row.F_t_wc_Rd*1.0e-3,' [kN]', '(Column web in transverse tension)'])
                ws.append(['F_t_f_Rd = ', row.F_t_f_Rd*1.0e-3,' [kN]', '(Column flange in bending)'])
                ws.append(['F_T_1_Rd = ', row.F_T_Rdf[0]*1.0e-3,' [kN]'])
                ws.append(['F_T_2_Rd = ', row.F_T_Rdf[1]*1.0e-3,' [kN]'])
                ws.append(['F_T_3_Rd = ', row.F_T_Rdf[2]*1.0e-3,' [kN]'])
                ws.append(['F_T_1_2_Rd = ', row.F_T_Rdf[3]*1.0e-3,' [kN]'])
                ws.append(['Breaking mode of flange in bending: ', row.mode_f])

                ws.append(['F_t_wb_Rd = ', row.F_t_wb_Rd * 1.0e-3, ' [kN]', '(Beam web in tension)'])
                ws.append(['F_t_p_Rd = ', row.F_t_p_Rd * 1.0e-3, ' [kN]', '(End-plate in bending)'])
                ws.append(['F_T_1_Rd = ', row.F_T_Rdp[0] * 1.0e-3, ' [kN]'])
                ws.append(['F_T_2_Rd = ', row.F_T_Rdp[1] * 1.0e-3, ' [kN]'])
                ws.append(['F_T_3_Rd = ', row.F_T_Rdp[2] * 1.0e-3, ' [kN]'])
                ws.append(['F_T_1_2_Rd = ', row.F_T_Rdp[3] * 1.0e-3, ' [kN]'])
                ws.append(['Breaking mode of flange in bending: ', row.mode_f])
                ws.append(['Breaking mode of end-plate in bending: ',row.mode_p])
                ws.append([])

            ws.append([])
            ws.append(['Possible row groups: '])
            ws.append([str(sd.groups)])
            ws.append([])
            ws.append([])

            for j in range(0, len(sd.row_group)):
                group = sd.row_group[j]

                ws.append(['Row group ', group.id+1, 'p = ',str(group.p),' [mm]', str(group.group)])
                for row in group.row:
                    ws.append(['Row id = ', row.id, 'z = ', row.z,' [mm]'])
                ws.append(['l_eff_1f = ',group.l_eff_1f,' [mm]', 'l_eff_2f = ',group.l_eff_2f,' [mm]'])
                ws.append(['l_eff_1p = ',group.l_eff_1p,' [mm]', 'l_eff_2p = ',group.l_eff_2p,' [mm]'])

                ws.append(['Tension capacity of row group is defined by ', group.mode])
                ws.append(['Total tension capacity is ',group.F_t_Rd*1.0e-3,' [kN]'])

                ws.append(['F_t_wc_Rd = ',group.F_t_wc_Rd*1.0e-3,' [kN]', '(Column web in transverse tension)'])
                ws.append(['F_t_f_Rd = ',group.F_t_f_Rd*1.0e-3,' [kN]','(Column flange in bending)'])
                ws.append(['F_T_1_Rd = ', group.F_T_Rdf[0]*1.0e-3,' [kN]'])
                ws.append(['F_T_2_Rd = ', group.F_T_Rdf[1]*1.0e-3,' [kN]'])
                ws.append(['F_T_3_Rd = ', group.F_T_Rdf[2]*1.0e-3,' [kN]'])
                ws.append(['F_T_1_2_Rd = ', group.F_T_Rdf[3]*1.0e-3,' [kN]'])
                ws.append(['Breaking mode of flange in bending:',group.mode_f])

                ws.append(['F_t_wb_Rd = ', group.F_t_wb_Rd * 1.0e-3, ' [kN]', '(Beam web in tension)'])
                ws.append(['F_t_p_Rd = ', group.F_t_p_Rd * 1.0e-3, ' [kN]', '(End-plate in bending)'])
                ws.append(['F_T_1_Rd = ', group.F_T_Rdp[0] * 1.0e-3, ' [kN]'])
                ws.append(['F_T_2_Rd = ', group.F_T_Rdp[1] * 1.0e-3, ' [kN]'])
                ws.append(['F_T_3_Rd = ', group.F_T_Rdp[2] * 1.0e-3, ' [kN]'])
                ws.append(['F_T_1_2_Rd = ', group.F_T_Rdp[3] * 1.0e-3, ' [kN]'])
                ws.append(['Breaking mode of flange in bending:', group.mode_f])
                ws.append(['Breaking mode of end-plate in bending:', group.mode_p])
                ws.append([])

            # Saving plots
            #cn.print_geom(beam_id, side=True, front=True, savepath=path, save_only=True)
            #img_front = Image(path + 'front.png')
            #ws.add_image(img=img_front, anchor="K12")
            #img_side = Image(path + 'side.png')
            #ws.add_image(img=img_side, anchor="K82")

    ws.append(['V_wp_Ed/V_wp_Rd = ', cn.n_V_wp,' (Column web shear utilization rate)'])

