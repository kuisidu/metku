# Imported libraries
from tables_and_tuples import *
from classes import EndPlate
from timeit import default_timer as timer
from optimization import Optimizer
from math import sqrt
import matplotlib.pyplot as plt
from write_excel import write_to_excel
from write_excel import open_workbook
from openpyxl.chart import BarChart, Reference


for beam_size in ['IPE 200']:#, 'IPE 400', 'IPE 500', 'IPE 600']:

    column_config = "column end"            # "continuous column" or "column end"
    column_end_length = 0                   # Column length over highest point of end-plate [mm],
                                            # If column_config="continuous column" and given value column_end_length=0
                                            # value column_end_length=4500.0 is used in calculations
                                            # Not implemented yet: "none"(= beam to beam connection)
    sway = "sway"
    cost_function = 'Diaz'                  # Used cost function: 'Haapio' or 'Diaz'
    test_method = "VT"                      # Test method for final inspection of connection
                                            # VT = Visual testing, UT = Ultrasonic testing, MT = Magnetic particle testing
    paint = "alkyd"                         # Paint used in final painting of parts: alkyd, epoxy, polyurethane or acryl
    cn = EndPlate(column_config, column_end_length, sway, cost_function, test_method, paint)
    cn.add_column(size='HE 140 B', material="S275", L=4000)
    cn.add_column_load(side='over')     # Loading over connection
    cn.add_column_load(side='under')    # Loading under connection
    cn.add_beam(beam_id=0, size=beam_size, material="S275", L=6000)
    sd = cn.side[0]
    sd.beam.loading(My=22, Vz=26.325)
    b = sd.beam.b

    result = []

    run = -1
    evals = 0
    fun_evals = []

    margin = 0.05
    S_j_limit = 9000.0*1.0e6

    method='SLSQP'

    file_name = 'Test_Diaz_A_0,05_c_' + str(cn.column.size) + '_b_' + str(beam_size)
    path = 'C:\\Users\LaHie1\Documents\Dippa\Excel taulukot\Diaz\\'
    wb = open_workbook(path=path, file_name=file_name)

    #for paksuus in [1, 2, 3, 4, 5, 6]:
    for sd.d in [12]:
    #for sd.d in bolt_sizes:
        cn.add_end_plate(beam_id=0, b=b, t=15, upper_overhang=4.0*sd.d, material="S275", equ=True)
        sd.e = 1.2*sd.d
        for k in range(1, 2):
            start = timer()
            run += 1
            print("\nBolt size {0}, rows {1}:".format(sd.d, k + 1))
            sd.row = []     # Initializing rows to zero
            # Define optimization parameters
            result.append(Optimizer(cn, beam_id=0))
            result[run].d = sd.d
            result[run].rows = k + 1
            z = 0.0
            for i in range(0, result[run].rows):
                if i == 0:
                    if sd.upper_overhang > 2.0 * 1.2 * sd.d:
                        z = 1.5 * sd.d
                    else:
                        z = sd.upper_overhang + sd.beam.t_f + 0.8*sqrt(2.0)*sd.weld_f + 1.5*sd.d
                else:
                    z = max(z + 2.4*sd.d, sd.upper_overhang + sd.beam.t_f + 0.8*sqrt(2.0)*sd.weld_f + 2.4*sd.d)
                cn.add_row(beam_id=0, z=z)

            # Add objective
            #result[run].add_objective(fun=lambda: -sd.S_j/1.0e6)#/((sd.beam.E*sd.beam.I_y)/sd.beam.L))
            #result[run].add_objective(fun=lambda: -sd.M_j_Rd/10.0e6)
            result[run].add_objective(fun=lambda: cn.total_cost)

            # Add design variables
            def var_t_p(x): sd.plate_thickness = x  # Design variable 1, end-plate thickness
            result[run].add_design_variable(var=var_t_p, var_init=15.0, var_bounds=[5.0, 30.0],
                                            var_name="t_p")

            def var_e(x): sd.e = x  # Design variable 2, edge distance of bolts
            e_min = max(1.2*sd.d_0, 0.5*sd.d_w)
            e_max = min(0.5*(sd.plate_width - sd.beam.t_w) - max(0.8*sqrt(2)*sd.weld_w + 1.2*sd.d_0, sqrt(2)*sd.weld_w + 0.5*sd.d_w),
                        0.5*(cn.column.b - cn.column.t_w) - max(0.8*cn.column.r + 1.2*sd.d_0, cn.column.r + 0.5*sd.d_w))
            if e_min > e_max:
                print("e_min > e_max, {0:{fm}} > {1:{fm}}, end-plate not wide enough for bolt size {2}"
                      .format(e_min, e_max, sd.d, fm='.2f'))
                continue
            sd.e = (e_min + e_max)/2.0  # Edge distance
            result[run].add_design_variable(var=var_e, var_init=0.5*(e_min + e_max),
                                                     var_bounds=[e_min, e_max], var_name="e")

            z_min = 0.0
            z_max = 2000.0

            def var3(x): sd.row[0].z = x  # Design variable 3, z coord of 1st row, on upper overhang
            #z_min1 = max(1.2*sd.d_0, 0.5*sd.d_w)
            #z_max1 = sd.upper_overhang - max(0.8*sqrt(2.0)*sd.weld_f + 1.2*sd.d_0, sqrt(2.0)*sd.weld_f + 0.5*sd.d_w)
            #if z_min1 > z_max1:
            #    print("z_min1 > z_max1: {0:{fm}} > {1:{fm}}".format(z_min1, z_max1, fm='.2f'))
            #    continue
            result[run].add_design_variable(var=var3, var_init=sd.row[0].z,
                                                     var_bounds=[z_min, z_max], var_name="row[0].z")

            #z_min = sd.upper_overhang + sd.beam.t_f + max(0.8*sqrt(2.0)*sd.weld_f + 1.2*sd.d_0, sqrt(2.0)*sd.weld_f + 0.5*sd.d_w)
            #z_max = sd.plate_height - sd.lower_overhang - sd.beam.t_f - max(0.8*sqrt(2.0)*sd.weld_f + 1.2*sd.d_0, sqrt(2.0)*sd.weld_f + 0.5*sd.d_w)

            if result[run].rows > 1:
                def var4(x): sd.row[1].z = x  # Design variable 4, z coord of 1st row under tension flange
                result[run].add_design_variable(var=var4, var_init=sd.row[1].z,
                                                         var_bounds=[z_min, z_max], var_name="row[1].z")
            if result[run].rows > 2:
                def var_end(x): sd.row[-1].z = x  # Design variable 5, z coord of lowest tension row under flange
                result[run].add_design_variable(var=var_end, var_init=sd.row[-1].z,
                                                         var_bounds=[z_min, z_max],
                                                         var_name="row[-1].z")
            if result[run].rows > 3:
                def var5(x): sd.row[2].z = x  # Design variable 4, z coord of 2nd row under tension flange
                result[run].add_design_variable(var=var5, var_init=sd.row[2].z,
                                                         var_bounds=[z_min, z_max], var_name="row[2].z")

            #def h_up(x):
            #    h_up_temp = sd.upper_overhang
            #    sd.upper_overhang = x
            #    for row in sd.row:
            #        row.z = row.z - (h_up_temp - sd.upper_overhang)
            #h_up_min = max(0.8*sqrt(2.0)*sd.weld_f + 2.0*1.2*sd.d_0, sqrt(2.0)*sd.weld_f + sd.d_w)
            #result[run].add_design_variable(var=h_up, var_init=sd.upper_overhang,
            #                                var_bounds=[h_up_min, 1000.0], var_name="Upper overhang")

            # Add constraints
            S_j_min = (1.0 - margin)*S_j_limit
            S_j_max = (1.0 + margin)*S_j_limit
            constr1 = lambda: (sd.S_j - S_j_min) / S_j_min  # S_j >= S_j_min
            constr2 = lambda: (S_j_max - sd.S_j) / S_j_max  # S_j <= S_j_max
            M_j_limit = sd.beam.My  # [kNm]
            constr3 = lambda: (sd.M_j_Rd * 1.0e-6 - M_j_limit) / M_j_limit  # M_Rd >= M_j_limit
            constr4 = lambda: (sd.plate_width - 2.0*sd.e) - 2.4*sd.d_0  # w >= 2.4*d_0
            constr5 = lambda: 0.5*sd.plate_width - sd.e - 0.5*sd.beam.t_w - max(0.8*sqrt(2.0)*sd.weld_w + 1.2*sd.d_0, 0.5*sd.d_w)  # m >= 1.2*d_0
            constr6 = lambda: sd.row[0].z - max(1.2*sd.d_0, 0.5*sd.d_w)  # row[0].z >= 1.2*d_0
            constr7 = lambda: (sd.upper_overhang - max(0.8*sqrt(2.0)*sd.weld_f + 1.2*sd.d_0, sqrt(2.0)*sd.weld_f + 0.5*sd.d_w)) - sd.row[0].z  # row[0].z <= max_z

            local = 'global'
            result[run].add_constraint(type='ineq', fun=constr1, name="S_j >= S_j_min", var_update=local)
            result[run].add_constraint(type='ineq', fun=constr2, name="S_j <= S_j_max", var_update=local)
            result[run].add_constraint(type='ineq', fun=constr3, name="M_Rd >= M_j_limit", var_update=local)
            result[run].add_constraint(type='ineq', fun=constr4, name="w >= 2.4*d_0", var_update=local)
            result[run].add_constraint(type='ineq', fun=constr5, name="m >= 1.2*d_0", var_update=local)

            result[run].add_constraint(type='ineq', fun=constr6, name="row[0].z >= 1.2*d_0", var_update=local)
            result[run].add_constraint(type='ineq', fun=constr7, name="row[0].z <= max_z", var_update=local)
            if len(sd.row) > 1:
                constr8 = lambda: sd.row[1].z - (
                    sd.upper_overhang + sd.beam.t_f + max(0.8*sqrt(2.0)*sd.weld_f + 1.2*sd.d_0, sqrt(2.0)*sd.weld_f + 0.5*sd.d_w))       # row[1].z >= min_z
                result[run].add_constraint(type='ineq', fun=constr8, name="row[1].z >= min_z", var_update=local)
                constr_r1 = lambda: (sd.row[1].z - sd.row[0].z) - max(2.2*sd.d_0, sd.d_w)       # p01 >= 2.2*d_0
                result[run].add_constraint(type='ineq', fun=constr_r1, name="p01 >= 2.2*d_0", var_update=local)
            if len(sd.row) > 2:
                constr_r2 = lambda: (sd.row[2].z - sd.row[1].z) - max(2.2*sd.d_0, sd.d_w)       # p12 >= 2.2*d_0
                result[run].add_constraint(type='ineq', fun=constr_r2, name="p12 >= 2.2*d_0", var_update=local)
            if len(sd.row) > 3:
                constr_r3 = lambda: (sd.row[3].z - sd.row[2].z) - max(2.2*sd.d_0, sd.d_w)       # p23 >= 2.2*d_0
                result[run].add_constraint(type='ineq', fun=constr_r3, name="p23 >= 2.2*d_0", var_update=local)

            # Utilization rate of shear must be under 1.0
            result[run].add_constraint(type='ineq', fun=lambda: 1.0 - sd.n_V, name="n_V <= 1.0", var_update='global')

            # Call optimizer
            # 'SLSQP'=Sequential Least SQuares Programming
            # 'Diff_evolution'=Differential evolution (Can't give constraints!)
            try:
                result[run].solve(method=method)
            except:
                print('\nSomething went wrong in optimization run!\n')
                cn.info(values=True, cost_info=False, warnings_and_errors=False, figures=True, side=False, front=True)
                plt.pause(10.0)

            end = timer()
            print("\nElapsed time: " + str(end - start) + " [s]")
            print("Total evaluations = " + str(cn.evaluations))

            for const in result[run].constraints:
                print('Constraint {0} = {1}'.format(const['name'], const['fun'](result[run].res.x)))

            fun_evals.append(cn.evaluations - evals)
            evals = cn.evaluations
            print('Total function evaluations on optimization run: ' + str(fun_evals[-1]))
            print()
            result[run].info()
            #cn.info(values=False, cost_info=False, warnings_and_errors=False,
            #        figures=True, side=True, front=True)
            #plt.pause(4)

            result[run].update_design_variables_(result[run].res.x)

            # Print info of constraints
            print("\nConstraints:")
            print("S_j_min <= S_j <= S_j_max : {0:{fm}} <= {1:{fm}} <= {2:{fm}} [kNm/rad]"
                  .format((1.0 - margin) * S_j_limit * 1.0e-6, sd.S_j * 1.0e-6,
                          (1.0 + margin) * S_j_limit * 1.0e-6,
                          fm=".2f"))
            print("M_Rd >= M_min : {0:{fm}} >= {1:{fm}} [kNm]"
                  .format(sd.M_j_Rd * 1.0e-6, M_j_limit, fm=".2f"))
            print("w >= 2.4*d_0 : {0:{fm}} >= {1:{fm}} [mm]".format((sd.plate_width - 2.0*sd.e), 2.4*sd.d_0,
                                                                    fm=".2f"))
            print("m >= 1.2*d_0 : {0:{fm}} >= {1:{fm}} [mm]".format(
                (0.5*sd.plate_width - sd.e - 0.8*sd.weld_w - 0.5*sd.beam.t_w),
                1.2*sd.d_0, fm=".2f"))
            print("1.2*d_0 <= row[0].z <= z_max : {0:{fm}} <= {1:{fm}} <= {2:{fm}} [mm]"
                  .format(1.2*sd.d_0, sd.row[0].z, (sd.upper_overhang - 0.8*sd.weld_f - 1.2*sd.d_0), fm=".2f"))
            for i in range(1, len(sd.row)):
                print("p >= 2.2*d_0 : {0:{fm}} >= {1:{fm}} [mm]"
                      .format(sd.row[i].z - sd.row[i - 1].z, 2.2*sd.row[i].d_0, fm=".2f"))

            cn.info(values=False, cost_info=False, warnings_and_errors=False,
                    figures=True, side=True, front=True)

            sheet_name = 'M' + str(sd.d) + ', ' + str(k + 1) + ' rows'
            write_to_excel(wb=wb, cn=cn, sheet_name=sheet_name, time=str(end - start),
                           fun_evals=fun_evals, evals=evals, path=path)

            # Print info of constraints to excel
            ws = wb[sheet_name]
            ws.append([])
            ws.append(["Constraints:"])
            for const in result[run].constraints:
                ws.append([const['name'], const['fun'](result[run].res.x)])

            ws.append([])
            ws.append(['Design variable!','Current!','Inital!','Bounds'])
            for n in range(0, len(result[run].vars)):
                ws.append([result[run].var_name[n], result[run].res.x[n], result[run].x0[n], str(result[run].bnds[n])])
            ws.append(["\nInitial objective function value:", result[run].fun0])
            ws.append(["Final objective function value:", result[run].res.fun])

            ws = wb['Sheet']

            if result[run].res.message == "Optimization terminated successfully.":
                cell1 = 'B' + str(run+5)
                cell2 = 'C' + str(run+5)
                cell3 = 'D' + str(run+5)
                cell4 = 'E' + str(run+5)
                cell5 = 'F' + str(run+5)
                cell6 = 'G' + str(run+5)
                cell7 = 'H' + str(run+5)
                ws[cell1] = sheet_name
                ws[cell2] = sd.S_j*1.0e-6                                       # [kNm/rad]
                ws[cell3] = sd.S_j / ((sd.beam.E*sd.beam.I_y)/sd.beam.L)        # [-]
                ws[cell4] = cn.total_cost                                       # [e]
                ws[cell5] = sd.M_j_Rd*1.0e-6                                    # [kNm]
                ws[cell6] = fun_evals[-1]                                       # [kpl]
                ws[cell7] = end - start                                         # [s]

            fun_evals.append(cn.evaluations - evals)
            evals = cn.evaluations

            # Save the file
            wb.save(path + file_name + '.xlsx')

    # Sorting optimization results according to objective function value
    i = 0
    while i < len(result):
        if type(result[i].res).__name__ == 'NoneType':
            del result[i]
            i -= 1
        elif result[i].res.message != 'Optimization terminated successfully.':
            del result[i]
            i -= 1
        if len(result) == 0:
            break
        i += 1

    if len(result) == 0:
        print("No feasible results found!")
    else:
        result = sorted(result, key=lambda result: result.res.fun)

        print("\nBest using bolts M{0}, rows {1}: {2} [e]\n"
              .format(result[0].d, result[0].rows, result[0].res.fun))

        for val in result:
            if val is not None:
                print("\Connection stiffness, bolt M{0}, rows {1}: {2} [kNm/rad]"
                      .format(val.d, val.rows, val.res.fun))
                print("   t_p = {0:{fm}} [mm]".format(val.res.x[0], fm=".2f"))
                print("   e = {0:{fm}} [mm]".format(val.res.x[1], fm=".2f"))
                for i in range(0, len(val.res.x) - 2):
                    print("   row[{0}].z = {1:{fm}} [mm]".format(i, val.res.x[i + 2], fm=".2f"))

            if val == result[0]:
                val.cn.info(values=False, cost_info=False, warnings_and_errors=False,
                        figures=True, side=True, front=True)
                print(val.cn.side[0].row)

    ws = wb['Sheet']

    ws['B4'] = 'Name'
    ws['C4'] = 'Sj [kNm/rad]'
    ws['D4'] = 'Sj/(EI/L) [-]'
    ws['E4'] = 'Cost [e]'
    ws['F4'] = 'Mj.Rd [kNm]'

    ws['B2'] = '(EI/L) = '
    ws['C2'] = ((sd.beam.E*sd.beam.I_y)/sd.beam.L)*1.0e-6
    ws['D2'] = '[kNm/rad]'

    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 2
    chart1.title = "Kiertymisjäykkyys"
    chart1.y_axis.title = 'Sj [kNm/rad]'
    chart1.legend = None

    data = Reference(ws, min_col=3, min_row=5, max_row=len(result)+5, max_col=3)
    cats = Reference(ws, min_col=2, min_row=5, max_row=len(result)+5)
    chart1.add_data(data, titles_from_data=False)
    chart1.set_categories(cats)
    #chart1.shape = 1
    ws.add_chart(chart1, "H3")

    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 2
    chart2.title = "Suheellinen kiertymisjäykkyys"
    chart2.y_axis.title = 'Sj/(EI/L) [-]'
    chart2.legend = None

    data = Reference(ws, min_col=4, min_row=5, max_row=len(result) + 5, max_col=4)
    chart2.add_data(data, titles_from_data=False)
    chart2.set_categories(cats)
    #chart2.shape = 1
    ws.add_chart(chart2, "R3")

    chart3 = BarChart()
    chart3.type = "col"
    chart3.style = 2
    chart3.title = "Liitoksen kokonaiskustannukset"
    chart3.y_axis.title = '[e]'
    chart3.legend = None

    data = Reference(ws, min_col=5, min_row=5, max_row=len(result) + 5, max_col=5)
    chart3.add_data(data, titles_from_data=False)
    chart3.set_categories(cats)
    #chart3.shape = 1
    ws.add_chart(chart3, "R24")

    chart4 = BarChart()
    chart4.type = "col"
    chart4.style = 2
    chart4.title = "Momenttikapasiteetti"
    chart4.y_axis.title = 'Mj.Rd [kNm]'
    chart4.legend = None

    data = Reference(ws, min_col=6, min_row=5, max_row=len(result) + 5, max_col=6)
    chart4.add_data(data, titles_from_data=False)
    chart4.set_categories(cats)
    #chart4.shape = 1
    ws.add_chart(chart4, "H24")

    chart5 = BarChart()
    chart5.type = "col"
    chart5.style = 2
    chart5.title = "Liitoslaskentapohjan evaluointien määrä"
    chart5.y_axis.title = '[kpl]'
    chart5.legend = None

    data = Reference(ws, min_col=7, min_row=5, max_row=len(result) + 5, max_col=7)
    chart5.add_data(data, titles_from_data=False)
    chart5.set_categories(cats)
    ws.add_chart(chart5, "H44")

    chart6 = BarChart()
    chart6.type = "col"
    chart6.style = 2
    chart6.title = "Optimointiaika"
    chart6.y_axis.title = '[s]'
    chart6.legend = None

    data = Reference(ws, min_col=8, min_row=5, max_row=len(result) + 5, max_col=8)
    chart6.add_data(data, titles_from_data=False)
    chart6.set_categories(cats)
    ws.add_chart(chart6, "R44")

    # Save the file
    wb.save(path + file_name + '.xlsx')

    plt.show()
