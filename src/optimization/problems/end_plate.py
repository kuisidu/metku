"""
    Cost optimization of end plate connections
    
    @author: Kristo Mela
    @date: 31.1.2020
"""

import math
import numpy as np
from copy import deepcopy
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt

try:
    from src.frame2d.frame2d import SteelBeam
    from src.sections.steel import SteelSection
    from src.optimization.structopt import *
    from src.sections.steel.catalogue import ipe_profiles, shs_profiles, hea_profiles
    from src.sections.steel import ISection, WISection, RHS
except:
    #from optimization.structopt import     
    import optimization.structopt as sopt
    from structures.steel.plates import THICKNESSES
    from eurocodes.en1993.en1993_1_8.en1993_1_8 import SHEAR_ROW, ROW_OUTSIDE_BEAM_TENSION_FLANGE
    from eurocodes.en1993.en1993_1_8.en1993_1_8 import bolt_sizes

class EndPlateOpt(sopt.OptimizationProblem):
    """
    Class cost optimization of end plate connections

    """

    def __init__(self,
                 name="End-plate Optimization Problem",
                 structure=None,                 
                 constraints={'geometry':True},
                 variables={'bp':False, 'etop': False},
                 plate_thicknesses=THICKNESSES,
                 obj=None,
                 target_stiffness=None,
                 Sdev = 0.1):
        """ Constructor
            input:
                structure .. EndPlateJoint class object
                constraints .. list of constraints
                plate_thicknesses .. list of available plate thicknesses
                
        """
        super().__init__(name=name,
                         structure=structure)        
        self.plate_thicknesses = plate_thicknesses
        self.stiffness_dev = Sdev
        self.target_stiffness = target_stiffness
        self.Srigid = 0.0
        
        if variables is None:
            variables={'bp':False, 'etop': False}
            
        self.create_variables(variables)
        self._variables = variables
        
        
        if constraints is None:
            constraints = {
                'geometry': True,
                'bending_resistance': True,
                'shear_resistance': True,
                'target_stiffness': True,
                'rigid_joint': False,
                'pinned_joint': False,
                'initial_stiffness':False
            }
        
        
        
        if target_stiffness is None:
            constraints["target_stiffness"] = False
                        
        self.constraints = constraints
        self.create_constraints(**self.constraints)

        if obj is not None:
            self.create_objective(obj)
        elif obj is None:
            self.create_objective('cost')

            

    @property
    def available_constraints(self):
        """
        Returns all available constraints and their corresponding keyword
        :return:
        """
        return dict(geometry=False,
                    bending_resistance=False,
                    shear_resistance=False,
                    target_stiffness=False,
                    rigid_joint=False,
                    pinned_joint=False,
                    initial_stiffness=False)

    def vertical_row_distance_constraints(self, i):
        """
        Creates a constraint that states that two adjacent
        bolt rows need to be at least 2.2*d0 apart from each other
        vertically.
        :param i: index of the row
        :return: constraint functions in a dict
        """

        con_funs = {}
        d0 = self.structure.bolt_rows[0].bolt.d0
        """ Condition: z[i+1]-z[i] >= 2.2*d0 """
        rhs = 2.2*d0
        a = np.zeros(self.nvars())
        
        for j, var in enumerate(self.vars):
            if var.name == ('z' + str(i)):
                a[j] = 1
            elif var.name == ('z' + str(i+1)):
                a[j] = -1

        def p(x):
            self.substitute_variables(x)
            return self.structure.bolt_rows[i].z - self.structure.bolt_rows[i+1].z - 2.2*d0
            #return self.structure.bolt_rows[i+1].z_top - self.structure.bolt_rows[i].z_top - 2.2*d0

        #con_funs = {f'Row {i+2} - Row{i+1} >= 2.2*d0': p}
              
        con_name = f'Row {i+1} - Row{i} >= 2.2*d0'
        #con_fun = p        
        #return con_name, con_fun
        return sopt.LinearConstraint(a,rhs,'>',name=con_name)
    
    def extension_edge_distance(self):
        """ Limits the location of the row above the top flange of the beam
            etop + 0.5*hb - z1 >= 1.2*d0 or
            etop - z1 >= 1.2*d0 - 0.5*hb
        """
    
        a = np.zeros(self.nvars())
    
        for i, var in enumerate(self.vars):
            if var.name == 'eTop':
                a[i] = 1
            elif var.name == 'z1':
                a[i] = -1
                
    
        d0 = self.structure.bolt_rows[0].bolt.d0
        rhs = 1.2*d0 - 0.5*self.structure.beam.h
    
        return sopt.LinearConstraint(a,rhs,'>',name='Extension edge distance')
    
    def plate_edge_distance(self):
        """ Limits the location of the rows from the edge of end plate
            0.5*bp - x >= 1.2*d0
            x + 0.5*dw <= 0.5*bp
            ->
            0.5*bp - x >= max(1.2*d0,0.5*bp)
        """
    
        a = np.zeros(self.nvars())
    
        for i, var in enumerate(self.vars):
            if var.name == 'bp':
                a[i] = 0.5
            elif var.name == 'x':
                a[i] = -1
                
    
        d0 = self.structure.bolt_rows[0].bolt.d0
        dw = self.structure.bolt_rows[0].bolt.dw
        rhs = max(1.2*d0,0.5*dw)
    
        return sopt.LinearConstraint(a,rhs,'>',name='End plate edge distance')
    
    
    def moment_resistance_constraints(self):
        """ Moment resistance constraint:
            MjEd <= MjRd(x)
        """
        def Mcon(x):
            self.substitute_variables(x)
            #return 1-self.structure.MjRd*1e-6/self.structure.MjEd
            #return self.structure.MjEd-self.structure.MjRd*1e-6
            return self.structure.MjEd/(self.structure.MjRd*1e-6)-1
        
        con_name = "1-MjRd/MjEd <= 0"
        con_fun = Mcon
        
        return con_name, con_fun

    def shear_resistance_constraints(self):
        """ Moment resitance constraint:
            VjEd <= VjRd(x)
        """
        def Vcon(x):
            self.substitute_variables(x)
            #return 1-self.structure.VjRd*1e-3/self.structure.VjEd
            return self.structure.VjEd-self.structure.VjRd*1e-3
            #return self.structure.VjEd/(self.structure.VjRd*1e-3)-1
        
        con_name = "1-VjRd/VjEd <= 0"
        con_fun = Vcon
        
        return con_name, con_fun
    
    def stiffness_constraint(self,initial=False):
        """ Requirement for target stiffness
            input:
                S_target .. target stiffness (kNm)
                p .. allowed deviation from target stiffness (percent)
                
                Sj(x) <= (1+p)*S_target
                Sj(x) >= (1-p)*S_target
        """

        print(initial)

        def Scon_ub(x):
            """ Sj(x) <= (1+p)*S_target """
            self.substitute_variables(x)
            #return self.structure.rotational_stiffness()*1e-6/((1+self.stiffness_dev*1e-2)*self.target_stiffness) - 1
            if initial:
                return 1e-2*(self.structure.Sj_ini()*1e-6-(1+self.stiffness_dev*1e-2)*self.target_stiffness)
            else:
                return 1e-2*(self.structure.rotational_stiffness()*1e-6-(1+self.stiffness_dev*1e-2)*self.target_stiffness)

        def Scon_lb(x):
            """ Sj(x) >= (1-p)*S_target """
            self.substitute_variables(x)
            if initial:
                return 1e-2*(self.structure.Sj_ini()*1e-6-(1-self.stiffness_dev*1e-2)*self.target_stiffness)
            else:
            #return self.structure.rotational_stiffness()*1e-6/((1-self.stiffness_dev*1e-2)*self.target_stiffness) - 1
                return 1e-2*(self.structure.rotational_stiffness()*1e-6-(1-self.stiffness_dev*1e-2)*self.target_stiffness)
        
        cons = {}
        cons["Stiffness upper bound"] = Scon_ub
        cons["Stiffness lower bound"] = Scon_lb
        
        return cons
    
    def rigid_joint_constraint(self):
        """ Requirement that threshold initial rotational stiffness is exceeded
            so that the joint can be regarded as rigid
            input:
                S_rigid .. limit stiffness (kNm)
                
                S_j_init(x) >= S_rigid
                
        """

        def Scon_rigid(x):
            """ Sj_ini(x) >= S_rigid """
            self.substitute_variables(x)
            #return self.structure.rotational_stiffness()*1e-6/((1-self.stiffness_dev*1e-2)*self.target_stiffness) - 1
            return 1e0*(self.structure.Sj_ini()*1e-6-self.Srigid)
        
        con_name = "Stiffness for rigid joint"
        con_fun = Scon_rigid        
        
        return con_name, con_fun
    
    
    def create_constraints(self,
                           geometry=False,
                           bending_resistance=False,
                           shear_resistance=False,
                           target_stiffness=False,
                           rigid_joint=False,
                           pinned_joint=False,
                           initial_stiffness=False,
                           ):
        """
        Creates constraints

        NOTE! The name of these kwargs must match the key in the cons -dict
        created in constraint function creator method
        e.g.
        CORRECT:
        cons = {'compression': comp_fun, 'tension': ten_fun}
        FALSE:
        cons = {'comp_func': comp_fun, 'tension_con': ten_fun}

        :param geometry:        
        :return:
        """
        self.cons = []
        
        
        if geometry:
            for i in range(1,self.structure.nrows):
                if self.structure.bolt_rows[i].type != SHEAR_ROW:
                    con = self.vertical_row_distance_constraints(i)
                    self.add(con)
                """
                p_name, p_fun = self.vertical_row_distance_constraints(i)
                con = sopt.NonLinearConstraint(name=p_name,
                                               con_fun=p_fun,
                                               con_type='>',
                                               fea_required=False)                
                            #vars=mem)
                """
                    
                
            if self._variables['etop']:
                con = self.extension_edge_distance()
                self.add(con)
            
            if self._variables['bp']:
                con = self.plate_edge_distance()
                self.add(con)
            
        
        if bending_resistance:
            m_name, m_fun = self.moment_resistance_constraints()
            mcon = sopt.NonLinearConstraint(name=m_name,
                                            con_fun=m_fun,
                                            con_type='<',
                                            fea_required=True)
            self.add(mcon)

        if shear_resistance:
            v_name, v_fun = self.shear_resistance_constraints()
            mcon = sopt.NonLinearConstraint(name=v_name,
                                            con_fun=v_fun,
                                            con_type='<',
                                            fea_required=True)
            self.add(mcon)
            
        if target_stiffness:
            Scons = self.stiffness_constraint(initial_stiffness)
            
            for key, value in Scons.items():
                if key.find('upper') != -1:
                    """ upper bound """
                    ctype = '<'
                else:
                    """ lower bound """
                    ctype = '>'
                self.add(sopt.NonLinearConstraint(name=key,
                                                     con_fun=value,
                                                     con_type=ctype,
                                                     fea_required=True)
                         )

        if rigid_joint:
            r_name, r_fun = self.rigid_joint_constraint()
            
            self.add(sopt.NonLinearConstraint(name=r_name,
                                              con_fun=r_fun,
                                              con_type='>',
                                              fea_required=True)
                     )
            

    def create_variables(self,variables={'bp':False, 'etop': False}):
        """
        Creates design variables
        
        By default, the design variables are:
            tp .. end plate thickness
            x .. horizontal distance of all bolts from center axis of the beam
            zi .. vertical distance of bolt row i from top edge of the plate
            bp .. width of the end plate
            
        Alternatively, the zi could be replaced by:
            z1 .. distance of the top row from the top of the plate
            pi .. distance of adjacent rows
            -> this could be set to all rows such that the distance between
               selected rows would be equal. If zi variables are used,
               equality constraints need to be added to control the distance
               between rows.
        
        """
        self.vars = []

        """ End plate thickness """
        tp_var = sopt.Variable("tp",
                               lb=min(THICKNESSES),
                               ub=max(THICKNESSES),
                               target={"property":"t","objects":[self.structure.end_plate]},
                               value=self.structure.tp)
        
        self.add(tp_var)
        
        """ Distance from bolt rows from axis of the beam """
        d0 = self.structure.bolt_rows[0].bolt.d0
        dw = self.structure.bolt_rows[0].bolt.dw
        hb = self.structure.beam.h
        bb = self.structure.beam.b
        tb = self.structure.beam.tw
        tbf = self.structure.beam.tf
        bc = self.structure.col.b
        tc = self.structure.col.tw
        rc = self.structure.col.r
        af = self.structure.weld_f
        aw = self.structure.weld_w
        x_min = max(0.5*tb + math.sqrt(2)*aw + 0.5*dw,
                    0.5*tc + rc + 0.5*dw,
                    1.2*d0)
        # Distance from column flange edge
        x_max = min(0.5*bc-1.2*d0,
                    0.5*bc-0.5*dw)
        
        axis_x_var = sopt.Variable("x",
                                   lb=x_min,
                                   ub=x_max,
                                   target={"property":"xbolts","objects":[self.structure]},
                                   value=0.5*self.structure.bp-self.structure.ebolts)
        
        self.add(axis_x_var)
        
        """ Vertical distance of bolt rows 
        """
        if variables['etop']:
            """ If the top extension of the end plate is a variable,
                take the extension limit equal to 5*d0.
            """
            etop_ub = 5*d0
            lb = math.ceil(0.5*dw + math.sqrt(2)*af + 1.2*d0)
            etop_var = sopt.Variable("eTop",lb=lb, ub=etop_ub, target={"property":'etop','objects':[self.structure]},
                                  value=self.structure.etop)
            self.add(etop_var)
        else:
            """ If the top extension is not a variable, take the initial
                extension as the limit
            """
            etop_ub = self.structure.etop
            lup = self.structure.etop
        
        if variables['bp']:
            """ Create variable for end plate width """
            #bp_lb = bb + math.ceil(2*math.sqrt(2)*af)
            bp_lb = bb
            bp_ub = bc + 100.0
            bp_var = sopt.Variable("bp",lb=bp_lb, ub=bp_ub, target={"property":'bp','objects':[self.structure]},
                                  value=self.structure.etop)
        
            self.add(bp_var)
        # zlb = max(1.2*d0,0.5*dw)
        # zub = hb
        # lb = zlb
        # Bounds for the first row, above the top flange of the beam
        zlb = math.ceil(0.5*hb + math.sqrt(2)*af + 0.5*dw)
        zub = 0.5*hb + etop_ub
        #zub = hb
        #lb = zlb
        self.add(sopt.Variable("z1",lb=zlb, ub=zub, target={"property":'z','objects':[self.structure.bolt_rows[0]]},
                                  value=self.structure.bolt_rows[0].z))
        
        zub = math.floor(0.5*hb - tbf - math.sqrt(2)*af - 0.5*dw)
        zlb = -zub
        
        for i, row in enumerate(self.structure.bolt_rows):
            if row.type is not SHEAR_ROW and row.loc_end_plate is not ROW_OUTSIDE_BEAM_TENSION_FLANGE:            
                self.add(sopt.Variable("z"+f"{i+1}",lb=zlb, ub=zub, target={"property":'z','objects':[row]},
                                  value=row.z))                
                zub -= 2.2*d0
        
        """
        for i, row in enumerate(self.structure.bolt_rows):
            if row.type is not SHEAR_ROW and loc_end_plate is not ROW_OUTSIDE_BEAM_TENSION_FLANGE:
                if i == 0:
                    "" First row above tension flange (assume extended end plate
                        connection.)
                    ""
                    ub = min(zub,math.floor(lup-math.sqrt(2)*af-0.5*dw))
                    #lb = zlb
                elif i == 1:
                    ub = zub
                    lb = max(zlb,math.ceil(lup+math.sqrt(2)*af+0.5*dw+tb))
                else:
                    lb += 2.2*d0
                    
                self.add(sopt.Variable("z"+f"{i+1}",lb=lb, ub=ub, target={"property":'z_top','objects':[row]},
                                  value=0.5*hb+self.structure.etop-row.z))
        """
        
    def create_objective(self,name='cost'):

        def obj_fun_cost(x):
            self.substitute_variables(x)            
            return self.structure.cost()
        
        def obj_fun_stiffness(x):
            self.substitute_variables(x)            
            return self.structure.Sj_ini()*1e-6
        
        def obj_fun_moment(x):
            self.substitute_variables(x)            
            return self.structure.bending_resistance()*1e-6
        
        if name == 'cost':
            fun = obj_fun_cost
            sense = 'MIN'
        elif name == 'stiffness':
            fun = obj_fun_stiffness
            sense = 'MAX'
        elif name == 'moment':
            fun = obj_fun_moment
            sense = 'MAX'

        obj = sopt.ObjectiveFunction(name=name.capitalize(),
                                     obj_fun=fun,
                                     obj_type=sense,
                                     fea_required=True)
        self.add(obj)


def optimize_for_target_stiffness(conn,Starget,rho=1.0,initial=False):
    
            
    print(initial)
    # Create problem
    problem = EndPlateOpt(name="Minimum cost, target stiffness",
                          structure=conn,                                
                          constraints={
                                    'geometry': True,
                                    'bending_resistance': True ,
                                    'shear_resistance': True,
                                    'target_stiffness': True,
                                    'initial_stiffness': initial,
                                },
                          variables={'bp':True, 'etop': True},
                          target_stiffness=Starget,
                          Sdev=rho)        
    
    return problem

def optimize_for_stiffness(conn,problem_type='MAX'):
                
    # Create problem
    if problem_type == "MAX":
        pname= "Stiffness maximization"        
    else:
        pname= "Stiffness minimization"        
        
    
    problem = EndPlateOpt(name="Stiffness maximization",
                          structure=conn,                          
                          constraints={
                                    'geometry': True,
                                    'bending_resistance': False ,
                                    'shear_resistance': False,
                                    'target_stiffness': False,
                                    'initial_stiffness': False,
                                },
                          obj='stiffness',                     
                          variables={'bp':True, 'etop': True})    

    if problem_type == 'MIN':
        problem.obj.obj_type = "MIN"
    
    return problem

def optimize_for_max_moment_resistance(conn):
                
    # Create problem
    problem = EndPlateOpt(name="Moment resistance maximization",
                          structure=conn,                          
                          constraints={
                                    'geometry': True,
                                    'bending_resistance': False ,
                                    'shear_resistance': False,
                                    'target_stiffness': False,
                                    'initial_stiffness': False,
                                },
                          obj='moment',                     
                          variables={'bp':True, 'etop': True})        
    
    return problem

def optimize_for_rigid_joint(conn,Srigid):

    problem = EndPlateOpt(name="Minimum cost, rigid joint",
                          structure=conn,                                
                          constraints={
                                    'geometry': True,
                                    'bending_resistance': True ,
                                    'shear_resistance': True,
                                    'rigid_joint': True,
                                    'target_stiffness': False,
                                    'initial_stiffness': True,
                                },
                          variables={'bp':True, 'etop': True},
                          target_stiffness=Srigid,
                          Sdev=0.1)
    
    problem.Srigid = Srigid
    
    return problem

def write_excel(problem,x_best,f_best,filename="Results.xlsx"):
    
    def print_rows():
        for row in sheet.iter_rows(values_only=True):
            print(row)
    
    #wb = Workbook()
    try:
        wb = load_workbook(filename=filename)
        sheet = wb.active
    except:
        wb = Workbook()
        sheet = wb.active
        
        # Create column names
        col_names = ['Rows','Bolt size', problem.obj.name]
        col_names += [var.name for var in problem.vars]
        if problem.constraints['target_stiffness']:
            if problem.constraints['initial_stiffness']:
                col_names += ['Sj_ini']
            else:
                col_names += ['Sj']
        else:
            col_names += ['Sj,ini']
        col_names += ['MjRd']
        col_names += ['VjRd']
        sheet.append(col_names)        
    
    var_vals = [x for x in x_best]
    problem(x_best)
    col_vals = [problem.structure.nrows,problem.structure.bolt_size,
                f_best]
    col_vals += var_vals
    if problem.constraints['target_stiffness']:
        if problem.constraints['initial_stiffness']:
            Sj = problem.structure.Sj_ini()*1e-6
        else:
            Sj = problem.structure.rotational_stiffness()*1e-6
        col_vals += [Sj,
                     problem.structure.MjRd*1e-6,
                     problem.structure.VjRd*1e-3]        
    else:
        col_vals += [problem.structure.Sj_ini()*1e-6,
                     problem.structure.MjRd*1e-6,
                     problem.structure.VjRd*1e-3]
    
    #sheet.append(var_names)
    
    sheet.append(col_vals)
    #print_rows()
    
    wb.save(filename=filename)
    
    
    



def run_optimization_series(conn,Starget,problem_type='rigid_joint',
                            initial=False):
    """ Runs a series of optimization problems for the joint 'conn'
        Varied parameters:
            i) bolt size
            ii) bolt grade
            iii) number of bolt rows
    
    """
    
    res = []    
    
    res_dir = 'P:\\INTRA_HOME\\Tutkimus\\IXConnections\\'
    #filename= res_dir + 'Diaz_Joint_Results.xlsx'
    
    filename= res_dir + problem_type.capitalize() + '_Diaz.xlsx'
    #filename= res_dir + 'Max_Stiffness_Results.xlsx'
    #fig_name_base = 'RigidJoint_M'
    fig_dir = 'P:\\INTRA_HOME\\Tutkimus\\IXConnections\\figures\\'
    #fig_name_base = 'DiazJoint_M'
    #fig_name_base = 'Max_Stiffness_M'
    fig_name_base = problem_type.capitalize() + '_Diaz_M'
    
    
    for d in reversed([12, 16, 20, 24]): #bolt_sizes:
        print("**** COST OPTIMIZATION FOR BOLT DIAMETER {0:2.0f}".format(d))
        
        conn.bolt_size = d
        if problem_type == 'rigid_joint':
            problem = optimize_for_rigid_joint(conn, Starget)
        elif problem_type == 'target_stiffness':
            problem = optimize_for_target_stiffness(conn, Starget, initial=initial)
        elif problem_type == 'max_stiffness':
            problem = optimize_for_stiffness(conn,'MAX')
        elif problem_type == 'min_stiffness':
            problem = optimize_for_stiffness(conn,'MIN')
        elif problem_type == 'max_moment':
            problem = optimize_for_max_moment_resistance(conn)
        else:
            print("Warning: unrecognized problem type!")
        
        
        solver = TrustRegionConstr()
        x0 = [var.value for var in problem.vars]
        
        if problem_type == "min_stiffness":
            x0[0] = problem.vars[0].lb
        
        
        f_best, x_best, nit = solver.solve(problem, maxiter=500, x0=x0)
        problem.num_iters = nit
        
        problem(x_best, prec=5)
        
        fig_name = fig_name_base + str(d) + '_' + str(conn.nrows) + '_Rows.svg'
        #conn.draw(fig_dir + fig_name)
        #plt.close()
        res.append([d,problem,f_best,x_best])
        # Write results to Excel file
        #write_excel(problem,x_best,f_best,filename)

    # Finally, write initial parameters to Excel
    """
    wb = load_workbook(filename)
    sheet = wb.active
    
    sheet.insert_rows(idx=1,amount=6)
    if problem_type == 'rigid_joint':
        sheet["A1"] = 'Rigid joint cost minimization'
    else:
        sheet["A1"] = 'Cost minimization for target stiffness'
    sheet["A2"] = 'SjRigid'
    sheet["B2"] = problem.target_stiffness
    sheet["C2"] = 'kNm/rad'
    sheet["A3"] = 'MjEd'
    sheet["B3"] = problem.structure.MjEd
    sheet["C3"] = 'kNm'
    sheet["A4"] = 'VjEd'
    sheet["B4"] = problem.structure.VjEd
    sheet["C4"] = 'kN'
    sheet["A5"] = 'Column'
    sheet["B5"] = problem.structure.col.__repr__() + ' ' + problem.structure.col.material.__repr__()
    sheet["A6"] = 'Beam'
    sheet["B6"] = problem.structure.beam.__repr__() + ' ' + problem.structure.beam.material.__repr__()   

    wb.save(filename=filename)
    """

    return res

if __name__ == '__main__':
    
    from eurocodes.en1993.en1993_1_8 import en1993_1_8

    import structures.steel.end_plate_joint as ep
    from optimization.solvers.trust_region import TrustRegionConstr
    #from structures.steel.end_plate_joint import example_1
    
    #conn = ep.example_1()
    nrows = range(2,5)
    nrows = [4]
    
    for r in reversed(nrows):
        conn = ep.diax_ex_rows(MjEd=40.0,VjEd=32.0,rows=r)
        #conn.MjEd = 0.5*round(0.8*conn.beam.MRd[0]*1e-6)
        #conn.MjEd = 180
        #conn = ep.example_diaz(connection='C')
        Starget = 16000.0
        Lbeam =  6000
        #Lbeam = 12000
        Kb = conn.beam.E*conn.beam.Iy/Lbeam
        kb = 8
        Srigid = math.ceil(kb*Kb*1e-6)
        Starget = Srigid
        #Starget = 2.0*Kb*1e-6
        print("Target stiffness: {0:4.2f} kNm/rad".format(Starget))
    
        #conn.bending_resistance(True)
        #conn.rotational_stiffness(True)
        #conn.MjEd = 200.0 # 22.0
        #conn.VjEd = 200.0 # 26.325
        
        #conn.tp = 20
        res = run_optimization_series(conn,Starget,initial=False, problem_type='rigid_joint')
    
    #problem = optimize_for_rigid_joint(conn, Srigid)
    
    """
    problem = EndPlateOpt(name="TEST",
                          structure=conn,                                
                          constraints={
                                    'geometry': True,
                                    'bending_resistance': True ,
                                    'shear_resistance': True,
                                    'stiffness': True
                                },
                          variables={'bp':True, 'etop': True},
                          target_stiffness=46000.0,
                          Sdev=0.1)
    """
    
    """
    solver = TrustRegionConstr()
    x0 = [var.value for var in problem.vars]
    f_best, x_best, nit = solver.solve(problem, maxiter=400, x0=x0)
    problem.num_iters = nit
    problem(x_best, prec=5)
    """
    # solver = GA(pop_size=50, mut_rate=0.15)
    # x0 = [1 for var in problem.vars]
    # fopt, xopt = solver.solve(problem, x0=x0, maxiter=10, plot=True)
    # problem(xopt)
    # frame.plot_deflection(100)
